from openpyxl import load_workbook
import collections
# Заполняем эксель файл с 7 популярными браузерами и количеством посещений по месяцам
# и вставляем в эксель
def insert_data_task(row_to_isert,row_to_check,copare_dict,checking_dict,book_name,sheet_name,new_book_name):
    """ Функция для вставки данных по требуемому параметру.
    :param row_to_iser: значение номера начальной ячейки для вставки
    :param row_to_check: значение номера ячейки для проверки данных
    :param copare_dict: словарь в котором ищутся совпадения по популярным обьектам
    :param checking_dict: список популярных объектов для сравнения
    :param book_name: имя открываемого файла (прим.Книга1.xlsx)
    :param sheet_name: имя рабочего листа (прим.Лист1)
    :param new_book_name: имя нового файла, может совпадать со старым (прим.Книга2.xlsx) """
    row_cnt = row_to_isert
    row_cnt_check = row_to_check
    wb = load_workbook(filename=book_name)
    sheet = wb[sheet_name]
    for chek_el in checking_dict:
        for k, v in copare_dict.items():
            if chek_el[0] == k:
                for k1, v1 in v.items():
                    for j in range(3, 15):
                        if sheet.cell(row=row_cnt_check, column=j).value == k1:
                            sheet.cell(row=row_cnt, column=j).value = v1
                            sheet.cell(row=row_cnt, column=1).value = chek_el[0]
                            sheet.cell(row=row_cnt, column=2).value = chek_el[1]
        row_cnt += 1
    return(wb.save(filename=new_book_name))


# функция для подсеча итогов по месяцам и заполнения их в файл эксель
def total_count(start_column, start_row, column_border, row_border, book_name, sheet_name, new_book_name):
    """Подсчет итогов в эксель файле.
    :param start_column: начальный столбец отсчета
    :param start_row: начальная строка отсчета
    :param column_border: граница заполнения данных для столбцов
    :param row_border: граница заполнения данных для строк
    :param book_name: имя открываемого файла (прим.Книга1.xlsx)
    :param sheet_name: имя рабочего листа (прим.Лист1)
    :param new_book_name: имя нового файла, может совпадать со старым (прим.Книга2.xlsx)
    """
    wb = load_workbook(filename=book_name)
    sheet = wb[sheet_name]
    column_name = int(start_column)
    row_name = int(start_row)
    sum = 0
    for col in range(column_name, column_border):
        for r in range(row_name, row_border):
            number = sheet.cell(row=r, column=col).value
            sum += number
            sheet.cell(row=row_border, column=col).value = sum
        sum = 0
        column_name += 1
    return (wb.save(filename=new_book_name))


# Функция определения самого популярного или не популярного товара
# среди мужчин или женщин и заполнение этих данных в эксель
def preferences_customers(cheking_dict,gender, indicator, needed_inf, insert_adress, book_name, sheet_name, new_book_name):
    """Поиск товара и  заполнение его в файл
    :param gender: пол для анализа продаж
    :param indicator: определение самый популярный или непопулярный товар (имеет значение +/-)
    :param needed_inf: информация о наименовании или кол-ве продаж ( наименование 0, кол-во 1)
    :param insert_adress: адресс ячейки для вставки информации в ввиде A1 
    :param book_name: имя открываемого файла (прим.Книга1.xlsx)
    :param sheet_name: имя рабочего листа (прим.Лист1)
    :param new_book_name: имя нового файла, может совпадать со старым (прим.Книга2.xlsx)
    """
    preferece_of_m = []
    for preference in cheking_dict:
        for k, v in preference.items():
            if k == gender:
                for i in v:
                    add_el = i
                    preferece_of_m.append(add_el)
    product_counter = collections.Counter(preferece_of_m)
    len_product_counter = len(product_counter)
    if indicator == '+':
        most_common_product_m = product_counter.most_common()[: + (len_product_counter + 1): + 1][0][needed_inf]
    elif indicator == '-': 
        most_common_product_m = product_counter.most_common()[: - (len_product_counter + 1): - 1][0][needed_inf] 
    wb = load_workbook(filename=book_name)
    sheet = wb[sheet_name]
    sheet[insert_adress] = most_common_product_m
    return(wb.save(filename=new_book_name))    